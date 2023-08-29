from openpyxl import Workbook, load_workbook
import os

class logger():
    def __init__(self):
        self.log_directory=os.getcwd() + "/logs"
        self.wb = Workbook()
        self.ws = self.wb.active

    def set_log_file(self):
        self.current_log_file=self.log_directory



wb = load_workbook("sample.xlsx")
ws = wb.active
ws['A1'] = 43
ws.append([1, 2, 3])
wb.save("sample.xlsx")